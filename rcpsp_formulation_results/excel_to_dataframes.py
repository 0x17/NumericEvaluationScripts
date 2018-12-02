import openpyxl
import os
import pandas as pd


def result_files(path):
    return [f for f in os.listdir(path) if f.endswith('.xlsx') and f.startswith('RESULTS_')]


def name_from_pair(fn, par, inst):
    prefix = fn.replace('RESULTS_', '').replace('.xlsx', '')
    return prefix + str(par) + '_' + str(inst) + '.sm'


def to_val(row): return map(lambda x: x.value, row)


def df_for_worksheet(fn, ws):
    header = 'feasible;solvetime;gap'.split(';')
    data, index = [], []
    for row in ws.iter_rows(min_row=2, max_col=9, max_row=481):
        par, inst, obj, solvetime, _, lb, ub, best, gap = to_val(row)
        feasible = obj != 0
        index.append(name_from_pair(fn, par, inst))
        data.append([feasible, solvetime, gap])
    return pd.DataFrame(index=index, data=data, columns=header)


def results_from_excel(fn):
    with openpyxl.load_workbook(fn) as wb:
        model_names = wb.sheetnames
        return {model_name: df_for_worksheet(fn, wb[model_name]) for model_name in model_names}


def result_map(path):
    return {fn: results_from_excel(fn) for fn in result_files(path)}


def model_names(dfs):
    return list(next(iter(dfs.values())).keys())


def testset_names(dfs):
    return list(dfs.keys())


def df_for_worksheets(filenames, worksheets):
    header = 'feasible;solvetime;gap'.split(';')
    data, index = [], []
    for ws_ix, ws in enumerate(worksheets):
        is_rg = 'RanGen2' in filenames[ws_ix]
        is_j90 = 'j90' in filenames[ws_ix]
        rg_set_no = 0
        testset_size = 1800 if is_rg else 480
        for row in ws.iter_rows(min_row=2, max_col=9, max_row=testset_size+1):
            if is_rg:
                setname, inst, obj, solvetime, _, lb, ub, best, gap = to_val(row)
                if setname is not None and len(setname.strip())>0:
                    rg_set_no = int(setname.split()[1])
                gap = 0 if isinstance(gap, str) and gap == '-' else gap
                inst_name = f'RG30_Set{rg_set_no}_Pat{inst}.rcp'
            else:
                par, inst, obj, solvetime, _, lb, ub, best, gap = to_val(row)
                solvetime = min(solvetime, 1200) if is_j90 else min(solvetime, 600)
                inst_name = name_from_pair(filenames[ws_ix], par, inst)
            feasible = obj != 0
            index.append(inst_name)
            data.append([feasible, solvetime, gap])
    df = pd.DataFrame(index=index, data=data, columns=header)
    df.index.name = 'instance'
    return df


def merge_results_for_model(model_name):
    print(f'Merging data for model: {model_name}')
    workbooks, worksheets = [], []
    filenames = result_files('.')
    for rf in filenames:
        print(f'\tTaking data from {rf}')
        workbooks.append(openpyxl.load_workbook(rf))
        if model_name in workbooks[-1].sheetnames:
            worksheets.append(workbooks[-1][model_name])
        else:
            print(f'\tWarning: Missing {model_name} from results in {rf}!')
    df_for_worksheets(filenames, worksheets).to_csv(f'{model_name}.csv')
    for wb in workbooks:
        wb.close()


def merge_results_for_all_models():
    wb = openpyxl.load_workbook(result_files('.')[0])
    model_names = wb.sheetnames
    wb.close()

    for model_name in model_names:
        merge_results_for_model(model_name)


if __name__ == '__main__':
    # map = result_map('.')
    merge_results_for_all_models()
