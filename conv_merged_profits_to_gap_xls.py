import sys
import openpyxl
import os


def show_usage():
    print('Wrong number of arguments!')
    print('Usage: python conv_merged_profits_to_gap_xls.py [batch dir_prefix] merged_profit_file.txt')
    print('Expected CSV-format: instance;method1;...;methodN;methodRef')


def gaps_for_row(row):
    gaprow = []
    for i in range(1, len(row) - 1):
        gaprow.append((round(float(row[-1]), 4) - round(float(row[i]), 4)) / round(float(row[-1]), 4))
    return gaprow


class InputData:
    def __init__(self, fn):
        with open(fn) as fp:
            lines = list(map(lambda line: line.replace(',', '.'), fp.readlines()))
            self.header_row = lines[0].rstrip().split(';')
            self.full_mx = list(map(lambda line: line.rstrip().split(';'), lines[1:]))
            self.ref_col = list(map(lambda row: row[-1], self.full_mx))
            self.instance_col = list(map(lambda row: row[0], self.full_mx))
            self.profit_mx = list(map(lambda row: row[1:], self.full_mx))
            self.gap_mx = list(map(gaps_for_row, self.full_mx))


def translate_name(method_str):
    mapping = {
        'GA0': 'GA beta',
        'GA3': 'GA zr',
        'GA4': 'GA zrt',
        'LocalSolverNative0': 'LS beta',
        'LocalSolverNative3': 'LS zr',
        'LocalSolverNative4': 'LS zrt',
        'Results_': ' ',
        '_': ' ',
        '.txt': ''
    }
    s = method_str
    for pattern, replacement in mapping.items():
        s = s.replace(pattern, replacement)
    return s


def convert_to_xls(obj, out_fn):
    wb = openpyxl.Workbook()
    wb.guess_types = True

    row_offset, col_offset = 2, 2

    def add_frame(sheet, with_ref=False):
        for j in range(len(obj.header_row) - (0 if with_ref else 1)):
            hname = obj.header_row[j] if j == 0 else translate_name(obj.header_row[j])
            sheet.cell(row=row_offset - 1, column=col_offset - 1 + j, value=hname)
        for i in range(len(obj.instance_col)):
            sheet.cell(row=i + row_offset, column=col_offset - 1, value=obj.instance_col[i])

    psheet = wb.create_sheet('Profits')
    add_frame(psheet, True)

    for i in range(len(obj.profit_mx)):
        for j in range(len(obj.profit_mx[0])):
            psheet.cell(row=i + row_offset, column=j + col_offset, value=obj.profit_mx[i][j])

    gsheet = wb.create_sheet('Gaps')
    add_frame(gsheet)

    for i in range(len(obj.gap_mx)):
        for j in range(len(obj.gap_mx[0])):
            c = gsheet.cell(row=i + row_offset, column=j + col_offset, value=obj.gap_mx[i][j])
            c.number_format = '0.00%'

    def average_gap_for_method_ix(j):
        accum = 0
        ngaps = len(obj.gap_mx)
        for i in range(ngaps):
            accum += obj.gap_mx[i][j]
        return float(accum) / float(ngaps)

    asheet = wb.create_sheet('Avg. Gaps')
    for j in range(len(obj.header_row) - 2):
        asheet.cell(row=row_offset - 1, column=col_offset - 1 + j, value=translate_name(obj.header_row[j + 1]))
        c = asheet.cell(row=row_offset, column=col_offset - 1 + j, value=average_gap_for_method_ix(j))
        c.number_format = '0.00%'

    del wb['Sheet']

    wb.save(out_fn)


def convert_file(fn):
    assert (fn.endswith('.txt') or fn.endswith('.csv'))
    fn_xls = fn.split('.')[-2] + '.xls'
    convert_to_xls(InputData(fn), fn_xls)


def batch_convert(dir_prefix, fn):
    for dname in os.listdir('.'):
        if dname.startswith(dir_prefix) and os.path.isfile(dname + '/' + fn):
            convert_file(dname + '/' + fn)


if len(sys.argv) < 2:
    show_usage()
else:
    first_arg = sys.argv[1]
    if first_arg == 'batch':
        if len(sys.argv) != 4:
            show_usage()
        else:
            dir_prefix = sys.argv[2]
            batch_convert(dir_prefix, sys.argv[3])
    else:
        convert_file(first_arg)
